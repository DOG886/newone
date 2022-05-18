import xlrd
import xlsxwriter
import xlwt
import json
import requests
import openpyxl
from openpyxl import load_workbook

from Excel.install import i

wb = openpyxl.load_workbook('/Applications/办公自动化/Excel/cs.xlsx')
ws = wb.active
# ws.title = "测试"
# ws.sheet_properties.tabColor = "1072BA"
wb.sheetnames
# wa = wb[sheets[0]]
#
# print(wa)
# sheets = wb.sheetnames
# for sheet in sheets:
#     print(sheets)
# for sheet in wb:
#     print
#     sheet.title
# cole=ws['C']
# print(cole)
# for col in ws.iter_cols(min_row=1,max_col=3,max_row=2):
#     for cell in col:
#         print(cell)
# cellValue = ws.cell(row=1, column=1).value
sheet = wb['Sheet1']
# print(sheet.dimensions)
# # print(sheet["A1"].value,sheet["G18"].value)
# for i in sheet["A"]:
#     for j in i:
#         print(j.value)
    # print(sheet['B'])


def list1(cell):
    List=[]
    for ce in cell:
        List.append(ce.value)
    return List


idList = list1(sheet['B'])
# print(idList)


def get_location_in_list(x,target):
    step = -1
    items = list()
    for i in range(x.count(target)):
        y = x[step + 1:].index(target)
        step = step + y +1
        items.append(step)
    return items


idpos=get_location_in_list(idList,"name")
print(idpos)